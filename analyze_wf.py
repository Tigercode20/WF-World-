import openpyxl
import json

# Open workbook
wb = openpyxl.load_workbook(r'd:\tiger\Nutrition Manager\System wep\WF WORLD.xlsx', read_only=True)

result = {
    'sheets': wb.sheetnames,
    'sheet_details': {}
}

# Analyze each sheet
for sheet_name in wb.sheetnames[:15]:  # First 15 sheets
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(min_row=1, max_row=3, values_only=True))
    
    if rows:
        result['sheet_details'][sheet_name] = {
            'headers': list(rows[0]) if len(rows) > 0 else [],
            'row_count': ws.max_row,
            'sample_row': list(rows[1]) if len(rows) > 1 else []
        }

# Save to file
with open('wf_world_structure.json', 'w', encoding='utf-8') as f:
    json.dump(result, f, ensure_ascii=False, indent=2)

print(json.dumps(result, ensure_ascii=False, indent=2))
