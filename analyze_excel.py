import openpyxl
import json

# Open workbook
wb = openpyxl.load_workbook(r'd:\tiger\Nutrition Manager\System wep\WF World.xlsx', read_only=True)

result = {
    'sheets': wb.sheetnames[:10]
}

# Get Client_Registration headers
if 'Client_Registration' in wb.sheetnames:
    ws = wb['Client_Registration']
    rows = list(ws.iter_rows(min_row=1, max_row=5, values_only=True))
    if rows:
        headers = list(rows[0])
        result['Client_Registration_headers'] = headers
        if len(rows) > 1:
            result['Client_Registration_sample'] = dict(zip(headers, rows[1]))

# Get Sales headers
if 'Sales' in wb.sheetnames:
    ws = wb['Sales']
    rows = list(ws.iter_rows(min_row=1, max_row=5, values_only=True))
    if rows:
        headers = list(rows[0])
        result['Sales_headers'] = headers
        if len(rows) > 1:
            result['Sales_sample'] = dict(zip(headers, rows[1]))

# Save to file
with open('excel_structure.json', 'w', encoding='utf-8') as f:
    json.dump(result, f, ensure_ascii=False, indent=2)

print(json.dumps(result, ensure_ascii=False, indent=2))
