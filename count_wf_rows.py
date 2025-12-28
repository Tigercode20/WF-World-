import openpyxl

# Open workbook
wb = openpyxl.load_workbook(r'd:\tiger\Nutrition Manager\System wep\WF WORLD.xlsx', read_only=True)

print("Total sheets:", len(wb.sheetnames))
print()

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    # Get actual row count (not null)
    data_rows = 0
    for row in ws.iter_rows(min_row=1, max_row=100, values_only=True):
        if any(cell is not None and str(cell).strip() != '' for cell in row):
            data_rows += 1
    
    if data_rows > 0:
        print(f"Sheet: {sheet_name}")
        print(f"  Data rows (first 100): {data_rows}")
        print(f"  Max row: {ws.max_row}")
        print()
